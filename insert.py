from tkinter import *
import time,os,merge
import openpyxl as op
def inset(self):
    text=["Client Name","GSTIN","Amount","Date","Credit","Debit"]
    for i in range(len(self.va)):
        if self.va[i][0]!='':
            if str(i) in self.credit:
                self.va[i].append("True")
                self.va[i].append("False")
            else:
                self.va[i].append("False")
                self.va[i].append("True")
            d=self.va[i][0]+".xlsx"
            path="C:/Users/"+os.getlogin()+"/Desktop/GST Data/"+merge.init()+"/Accounts/"
            if os.path.exists(path):
                pass
            else:
                os.makedirs(path)
            if d in os.listdir(path=path):
                wb=op.load_workbook(path+d)
                ws=wb[self.va[i][0]]
                ws.append(self.va[i])
                wb.save(path+d)
            else:
                wb=op.Workbook()
                ws=wb.create_sheet()
                ws.title=self.va[i][0]
                ws.append(text)
                ws.append(self.va[i])
                wb.save(path+d)
class Put(Frame):
    def __init__(self,parent,sel):
        if sel!=None:
            sel.pack_forget()
        Frame.__init__(self,parent)
        self.pack(fill=BOTH,expand=True)
        frm=Frame(self)
        self.aman=[]
        self.credit,self.debit=[],[]
        frm.pack(expand=True)
        text=["Client Name","GSTIN","Credit","Debit","Amount","Date"]
        wid=["18","10","8","5","18","18"]
        self.va,self.bt=[],[]
        for i in range(6):
            Label(frm,text=text[i],width=wid[i]).pack(side=LEFT)
        for i in range(20):
            frm1=Frame(self)
            frm1.pack(expand=True)
            var=StringVar()
            var1=StringVar()
            var2=StringVar()
            self.va.append([frm1,var,var1,var2])
        for i in range(10):
            Entry(self.va[i][0],textvariable=self.va[i][1]).pack(side=LEFT)
            btn=Button(self.va[i][0],text=text[1])
            btn.pack(side=LEFT)
            btn.bind("<Button-1>",self.set)
            btn1=Button(self.va[i][0],text=text[2])
            btn1.pack(side=LEFT)
            btn1.bind("<Button-1>",self.set)
            self.bt.append([btn,btn1])
            Entry(self.va[i][0],textvariable=self.va[i][2]).pack(side=LEFT)
            e=Entry(self.va[i][0],textvariable=self.va[i][3])
            e.pack(side=LEFT)
            e.bind("<Button-1>",self.do)
        backbt=Button(self,text="Insert",command=lambda:inset(self))
        backbt.pack(fill=X,expand=True,side=BOTTOM)
        backbtn=Button(self,text="Back",command=lambda:(self.pack_forget(),sel.pack(fill=X)))
        backbtn.pack(fill=X,expand=True,side=BOTTOM)
    def do(self,event):
        event.widget.insert(0,time.strftime("%d/%m/%y"))
    def set(self,event):
        event.widget["state"]="disabled"
        for i in range(len(self.bt)):
            if self.bt[i][0]==event.widget:
                self.credit.append(str(i))
                self.bt[i][1]["state"]="disabled"
                break
            elif self.bt[i][1]==event.widget:
                self.debit.append(str(i))
                self.bt[i][0]["state"]="disabled"
                break
if __name__=="__main__":
    rott=Tk()
    pa=Put(rott,None)
    rott.title("Insert Data")
    rott.mainloop()
