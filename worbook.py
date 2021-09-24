from openpyxl import *
from tkinter import messagebox
import os,time,aman,insert,csvdata,gsthtml
class wordbook():
    def __init__(self):
        self.data=["M/s","Recipien GST No.","Invoice NO.","Date","Challan No.","Recipient P.O.S.","Sr.","Description of goods",
                   "Challan No","Vehicle No.","Owner/Transporter","HSN no","Quantity","Rate","Taxable Amount","CGST @","SGST @",
                   "IGST @","CGST","SGST","IGST","Grand Total","Rs in words","","CESS","E-way Bill NO.","Receiver Name","Loding NO.","Place of Loading"]
        #self.today=time.strftime("%b-%Y")+".xlsx"
    def insert(self,master):
        self.tt=str(master.tax[2])+(master.tax[0][:12]+str(master.gtotal)[:2])[:15]
        self.today=master.tax[3][3:]+".xlsx"
        path="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/"
        if self.today in os.listdir(path=path):
            wb=load_workbook(path+self.today)
            try:
                wb.remove(wb[self.tt])
                ws=wb.create_sheet()
                ws.title=self.tt
            except:
                ws=wb.create_sheet()
                ws.title=self.tt
        else:
            wb=Workbook()
            ws=wb.create_sheet()
            ws.title=self.tt
        ws.append(self.data)
        for i in master.data:
            ws.append(master.tax[:6]+i+master.tax[6:9]+[round(float(i[8])*master.tax[6]/100),round(float(i[8])*master.tax[7]/100),
                                                       round(float(i[8])*master.tax[8]/100),round(float(i[8])+float(master.cess)+
                                                        float(i[8])*master.tax[6]/100+float(i[8])*master.tax[7]/100+
                                                       float(i[8])*master.tax[8]/100),aman.amount(float(i[8])),"False",round(float(master.cess)),
                                                    i[3],master.reference,master.lrno,master.terms])
        ws.append(['','','','','','','','','','','','','','Total=',master.total,'','','',master.cgst,master.sgst,
                   master.igst,master.gtotal,master.tax[9]])
        wb.save(path+self.today)
        #c=csvdata.csvd()
        #c.csvfile(master)
        self.bill(master)
    def topdf(self,master,path):
        gsthtml.create(master)
        '''o=win32com.client.Dispatch("Excel.Application")
        o.Visible=False
        wb=o.Workbooks.Open(master)
        wb.ExportAsFixedFormat(0,path+master[48:-5]+".pdf")
        wb.close'''
    def bill(self,master):
        cwd=os.getcwd()
        wb=load_workbook(cwd+"/Data/maadurgacarrier.xlsx")
        ws=wb.active
        ws["A5"]=master.tax[0]
        ws["C7"]=master.tax[1]
        ws["E4"]=master.tax[2]
        ws["G4"]=master.tax[3]
        ws["F5"]=master.tax[4]
        ws["F7"]=master.tax[5]
        for row in range(9,len(master.data)+9):
            for col,j in [(1,0),(2,1),(4,5),(5,6),(6,7),(7,8)]:
                name=master.data[row-9]
                ws.cell(row=row,column=col,value="%s" %name[j])
        ws["G21"]=master.total
        ws["F22"]=master.tax[6]
        ws["G22"]=master.cgst
        ws["F23"]=master.tax[7]
        ws["G23"]=master.sgst
        ws["G24"]=master.igst
        ws["F24"]=master.tax[8]
        ws["G25"]=master.gtotal
        ws["A24"]=master.tax[9]
        tborder=styles.Border(top=styles.Side(style="hair",color="ff000000"))
        trborder=styles.Border(top=styles.Side(style="hair",color="00000000"),
                         right=styles.Side(style="hair",color="00000000"))
        rborder=styles.Border(right=styles.Side(style="hair",color="00000000"))
        lborder=styles.Border(left=styles.Side(style="hair",color="00000000"))
        dborder=styles.Border(bottom=styles.Side(style="hair",color="00000000"))
        
        for i in ["B1","D1","E1","B4","C23","B26","C26","C9"]:
            ws[i].border=tborder
        for i in ["G1","C4","D26","G26"]:
            ws[i].border=trborder
        for i in ["A25","A6","D5","D6"]:
            ws[i].border=lborder
        for i in ["G2","G3","G5","G6","G7","D29","G27","G28","G29","D28","D27"]:
            ws[i].border=rborder
        for i in ["B30","C30","E30","F30"]:
            ws[i].border=dborder
        ws["D8"].border=styles.Border(left=styles.Side(style="hair",color="00000000"),bottom=styles.Side(style="hair",color="00000000"))
        ws["G30"].border=styles.Border(right=styles.Side(style="hair",color="ff000000"),
                                       bottom=styles.Side(style="hair",color="ff000000"))
        ws["D30"].border=styles.Border(right=styles.Side(style="hair",color="ff000000"),
                                       bottom=styles.Side(style="hair",color="ff000000"))
        mains="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/excel/"+self.today[:-5]+"/"+self.tt+self.today
        try:
            os.mkdir("C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/excel/"+self.today[:-5])
        except:
            pass
        try:
            os.mkdir("C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/pdf/"+self.today[:-5])
        except:
            pass
        wb.save(mains)
        #mains="C:\\Users\\"+str(os.getlogin())+"\\Desktop\\GST Data\\excel\\"+self.today[:-5]+"\\"+self.tt+self.today
        self.topdf(master,"C:\\Users\\"+str(os.getlogin())+"\\Desktop\\GST Data\\pdf\\"+self.today[:-5]+"\\")
    def calgst(self,file):
        wb=load_workbook(file)
        gs=[]
        for i in wb:
            su=0
            if i.title[:4]=="Shee":
                del(i)
                continue
            for row in i:
                if row[13].value=="Total=":
                     su=su+int(row[18].value)+int(row[19].value)+int(row[20].value)
            gs.append(i.title)
            gs.append(su)
        file=open("C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/tax.txt","w+")
        for i in range(0,len(gs),2):
            file.write(time.strftime("%d-%b-%Y")+"\t"+gs[i]+"\t"+str(gs[i+1])+"\n")
        file.close()
        su=0
        file=open("C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/tax.txt","r")
        for line in file:
            su = su + int(line.split()[-1])
        file.close()
        messagebox.showinfo("True","Tax is "+str(su))
        print(su)
    def arngdata(self,file):
        path="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/data/"
        if os.path.exists(path):
            pass
        else:
            os.makedirs(path)
        wb=load_workbook(file)
        try:
            wb.remove(wb['Sheet'])
            wb.remove(wb['Sheet1'])
        except:
            pass
        for sheet in wb:
            i=0
            self.va=[]
            self.credit=[]
            if sheet.title[:12]+".xlsx" in os.listdir(path):
                wd=load_workbook(path+sheet.title[:12]+".xlsx")
                wp=wd[sheet.title[:10]]
            else:
                wd=Workbook()
                wp=wd.active
                wp.title=sheet.title[:10]
                wp.append(["M/s","Invoice No","Date","Description","Vehicle No.","Owner/Transporter","Quantity","Rate","Taxable Amount","Tax","Total Amount"])
            for row in sheet:
                if row[23].value=="False":
                    wp.append([row[0].value,row[2].value,row[3].value,row[7].value,row[9].value,row[10].value,row[12].value,row[13].value,row[14].value,int(row[21].value)-int(float(row[14].value)),row[21].value])
                    self.va.append([sheet["A2"].value,sheet["D2"].value,row[21].value])
                    self.credit.append(i)
                    i=i+1
                    print(i,row[6].value)
            try:
                wd.save(path+sheet.title[:12]+".xlsx")
            except Exception as err:
                print(str(err))
            try:
                insert.inset(self)
            except Exception as err:
                print("Error Happened"+str(err))
    def coalbil(self,master):
        self.today=master.tax[3][3:]+".xlsx"
        self.tt=str(master.tax[2])+(master.tax[0][:12]+str(master.gtotal)[:2])[:15]
        path="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/"
        if self.today in os.listdir(path=path):
            wb=load_workbook(path+self.today)
            try:
                wb.remove(wb[self.tt])
                ws=wb.create_sheet()
                ws.title=self.tt
            except:
                ws=wb.create_sheet()
                ws.title=self.tt
        else:
            wb=Workbook()
            ws=wb.create_sheet()
            ws.title=self.tt
        ws.append(self.data)
        for i in master.data:
            ws.append(master.tax[:6]+i+master.tax[6:9]+[round(float(i[8])*master.tax[6]/100),round(float(i[8])*master.tax[7]/100),
                                                       round(float(i[8])*master.tax[8]/100),round(float(i[8])+
                                                        float(i[8])*master.tax[6]/100+float(i[8])*master.tax[7]/100+
                                                       float(i[8])*master.tax[8]/100),aman.amount(float(i[8])),"False",master.cess,i[3],
                                                        master.reference,master.lrno,master.terms])
        ws.append(['','','','','','','','','','','','','','Total=',master.total,'','','',master.cgst,master.sgst,
                   master.igst,master.gtotal,master.tax[9],master.cess])
        wb.save(path+self.today)
        c=csvdata.csvd()
        c.csvfile(master)
        self.billed(master)
    def billed(self,master):
        cwd=os.getcwd()
        wb=load_workbook(cwd+"/Data/coalbill.xlsx")
        ws=wb.active
        ws["A5"]=master.tax[0]
        ws["C7"]=master.tax[1]
        ws["E4"]=master.tax[2]
        ws["G4"]=master.tax[3]
        ws["F5"]=master.tax[4]
        ws["F7"]=master.tax[5]
        ws["A9"]=master.data[0][0]
        ws["B9"]=master.data[0][1]
        ws["D9"]=master.data[0][5]
        ws["E9"]=master.data[0][6]
        ws["F9"]=master.data[0][7]
        ws["G9"]=master.data[0][8]
        ws["B10"]=master.data[0][3]
        ws["F11"]=master.tax[6]
        ws["G11"]=master.cgst
        ws["F12"]=master.tax[7]
        ws["G12"]=master.sgst
        ws["F13"]="400 MTS"
        ws["G13"]=master.cess
        ws["G14"]=master.gtotal
        ws["G15"]=master.gtotal
        ws["A17"]=master.tax[9]
        ws["A20"]=master.data[0][5]
        ws["C20"]=master.data[0][8]
        ws["D20"]=master.cgst
        ws["E20"]=master.sgst
        ws["F20"]=master.cess
        ws["G20"]=float(master.cgst)+float(master.sgst)+float(master.cess)
        tborder=styles.Border(top=styles.Side(style="hair",color="ff000000"))
        trborder=styles.Border(top=styles.Side(style="hair",color="00000000"),
                         right=styles.Side(style="hair",color="00000000"))
        brborder=styles.Border(bottom=styles.Side(style="hair",color="00000000"),
                         right=styles.Side(style="hair",color="00000000"))
        tlborder=styles.Border(top=styles.Side(style="hair",color="00000000"),
                         left=styles.Side(style="hair",color="00000000"))
        rborder=styles.Border(right=styles.Side(style="hair",color="00000000"))
        lborder=styles.Border(left=styles.Side(style="hair",color="00000000"))
        bborder=styles.Border(bottom=styles.Side(style="hair",color="00000000"))
        rlborder=styles.Border(right=styles.Side(style="hair",color="00000000"),
                         left=styles.Side(style="hair",color="00000000"))
        tbborder=styles.Border(top=styles.Side(style="hair",color="00000000"),
                         bottom=styles.Side(style="hair",color="00000000"))
        for i in ["B1","C1","D1","E1","F1","B4","C4","C9","C14","C15","B16","C16","C17","D27","D17","E17","F17","B17","B18","B20",
        "B22","C22","D22","F22","B23","C23",'B24','C24','B25','C25','B26','C26','B27','C27','E27','F27']:
            ws[i].border=tborder
        for i in ["G2","G3","G7","C6","G5","G6","B19","C19","G19","G16","G21","G23","G24","G25","B21","C21","D19",
        "D21","E19","E21","F19","F21"]:
            ws[i].border=rborder
        for i in ["G1","G17","G22","D23","D24","D25","D26"]:
            ws[i].border=trborder
        ws["A21"].border=ws["A19"].border=ws["A6"].border=lborder
        ws["G26"].border=brborder
        mains="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/coal/excel/"+self.tt+self.today
        wb.save(mains)
        self.topdff(master)
    def topdff(self,master):
        gsthtml.create(master)
        '''o=win32com.client.Dispatch("Excel.Application")
        o.Visible=False
        path="C:\\Users\\"+str(os.getlogin())+"\\Desktop\\GST Data\\Other\\pdf\\"
        wb=o.Workbooks.Open(master)
        wb.ExportAsFixedFormat(0,path+master[44:-5]+".pdf")
        wb.close'''
    def gst1(self,ma):
        file=time.strftime("%b-%Y")+".xlsx"
        path="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/"+file[:-5]+"/"
        if file in os.listdir(path):
            try:
                wb=load_workbook(path+file)
                ws=wb["b2b"]
            except:
                print("File cannot be opened")
        else:
            try:
                cwd=os.getcwd()
                wb=load_workbook(cwd+"/Data/GSTR1.xlsx")
                ws=wb["b2b"]
            except:
                print("File cannot be opened")
        ws.append([ma.tax[1],ma.tax[0],ma.tax[2],ma.tax[3],ma.gtotal,ma.tax[5],"N","","Regular","",ma.data[0][7],ma.data[0][8],ma.cess])
        wb.save(path+time.strftime("%b-%y")+".xlsx")
    def insertdata(self,master):
        self.ddata=["Seller","Seller GST No.","Invoice NO.","Date","Place of Supply","Sr. No","Description of goods","Challan No","GST @","CESS @","TCS/TDS @","HSN no",
                    "Quantity","Rate","Taxable Amount","CGST","SGST","IGST","CESS","TCS/TDS","Vehicle No.","Owner/Transporter","Grand Total","E-way Bill NO."]
        path="C:/Users/"+str(os.getlogin())+"/Desktop/GST Data/"+master.date.get()[3:]+"/"
        try:
            wb=load_workbook(path+master.date.get()[7:]+".xlsx")
        except:
            wb=Workbook()
        try:
            ws=wb[master.date.get()[3:6]]
        except:
            ws=wb.create_sheet()
            ws.title=master.date.get()[3:6]
            ws.append(self.ddata)
        total=0
        for i in range(len(master.data)):
            ws.append(master.tax+master.data[i]+master.gstt[i])
            a=master.gstt[i]
            total=float(master.data[i][-1])+a[0]+a[1]+a[2]+a[3]+a[4]
        ws.append(['','','','','','','','','','','','','','','','','','','','',master.vno.get(),master.owner.get(),total,master.eway.get()])
        wb.save(path+master.date.get()[7:]+".xlsx")
        try:
            if os.path.exists(path):
                wb=load_workbook(path+master.date.get()[3:]+".xlsx")
            else:
                os.mkdir(path)
        except:
            wb=Workbook()
        try:
            ws=wb["Input"]
        except:
            ws=wb.create_sheet()
            ws.title="Input"
            ws.append(["GSTIN of Supplier","Invoice Number","Invoice date","Invoice Value","Place Of Supply","Reverse Charge",
                       "Invoice Type","Rate","Taxable Value ","Integrated Tax Paid","Central Tax Paid","State/UT Tax Paid",
                       "Cess Paid","Eligibility For ITC","Availed ITC Integrated Tax","Availed ITC Central Tax",
                       "Availed ITC State/UT Tax","Availed ITC Cess"])
            for i in range(len(master.data)):
                a=master.gstt[i]
                ws.append(master.tax[1:4]+[total,"08BPXPS7371P1ZF",'','Regular',master.data[i][3],master.data[i][9],a[2],a[0],a[1],a[3],"Eligible",a[2],a[0],a[1],a[3]])
                wb.save(path+master.date.get()[3:]+".xlsx")
