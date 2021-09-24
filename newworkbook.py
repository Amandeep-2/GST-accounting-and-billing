import openpyxl as op
from tkinter import messagebox
import os,time,insert,csvdata,xmlfile,merge
#self.gst=[[cgst,sgst,igst,cess,tcs]]
#self.data=[[sr,desc,cessrate,tcsrate,gstrate,hsn no,quant,rate,unit,amount]]
#self.detail=[ms,msad1,msad2,pno,pgst]
#self.billdetail=[billno,date,chlno,eway,vno,revchrg,transport,lrno,gsttotalword]
class wordbook():
    def __init__(self):
        self.data=["M/s","Address line 1","Address line 2","Place of supply","Recipient GST No.","Invoice NO.","Date","Challan No.","E-way No.",
                   "Vehicle No.","Reverse Charge","Transporter Detail","Transport Document No.","CGST","SGST","IGST","Cess Amount","TCS/TDS Amount","Grand Total"]
        self.data2=["Sr. No.","Description of goods","Cess Rate","TCS/TDS","GST Rate","HSN No.","Quantity","Rate","per","Taxable Amount"]
        #self.today=time.strftime("%b-%Y")+".xlsx"
    def insert(self,master):
        self.tt=str(master.billdetail[0])+master.detail[0]
        self.today=master.billdetail[1][3:]+".xlsx"
        path=master.dirr+master.billdetail[1][3:]+"/"
        if os.path.exists(path+self.today):
            wb=op.load_workbook(path+self.today)
            try:
                wb.remove(wb[self.tt])
                ws=wb.create_sheet()
                ws.title=self.tt
            except:
                ws=wb.create_sheet()
                ws.title=self.tt
        else:
            wb=op.Workbook()
            ws=wb.create_sheet()
            ws.title=self.tt
        ws.append(self.data)
        ws.append(master.detail+master.billdetail[:-1]+[master.cgst,master.sgst,master.igst,master.cessu,master.tcst,master.gtotal])
        ws.append(self.data2)
        for i in range(len(master.data)):
            ws.append(master.data[i])
        wb.save(path+self.today)
    def calgst(self,fille):
        wb=op.load_workbook(fille)
        gs=[]
        for i in wb:
            su=0
            if i.title[:4]=="Shee":
                wb.remove(i)
                continue
            gs.append(i.title)
            try:
                gs.append(i[2][13].value)
                gs.append(i[2][14].value)
                gs.append(i[2][15].value)
                gs.append(i[2][16].value)
                gs.append(i[2][17].value)
            except:
                pass
        file1=open(os.path.dirname(fille)+"/tax.txt","w+")
        for i in range(0,len(gs),6):
            file1.write(time.strftime("%d-%b-%Y")+"\t"+gs[i]+"\t"+str(gs[i+1])+"\t"+str(gs[i+2])+"\t"+str(gs[i+3])+"\t"+str(gs[i+4])+"\t"+str(gs[i+5])+"\n")
        file1.close()
        cg=0
        sg=0
        ig=0
        ce=0
        tc=0
        file=open(os.path.dirname(fille)+"/tax.txt","r")
        for line in file:
            cg = cg + float(line.split()[-5])
            sg = sg + float(line.split()[-4])
            ig = ig + float(line.split()[-3])
            ce = ce + float(line.split()[-2])
            tc = tc + float(line.split()[-1])
        messagebox.showinfo("True","CGST is "+str(cg)+"\nSGST is "+str(sg)+"\nIGST is "+str(ig)+"\nCess is "+str(ce)+"\nTCS is "+str(tc))
        print("New workbook file :- ",cg,sg,ig,ce,tc)
        p=csvdata.csvd()
        p.csvfilecreate(fille)
    def arngdata(self,data):
        print(data)
        readdata=[]
        dirr="C:/Users/"+os.getlogin()+"/Desktop/GST Data/"+data+"/"
        path1=dirr+"temp/"
        path=dirr+"Data/"
        if os.path.exists(path):
            pass
        else:
            os.makedirs(path)
        if os.path.exists(path1):
            pass
        else:
            os.makedirs(path1)
        files=[]
        for ff,file,folder in os.walk(dirr):
            for i in folder:
                if ff[-4:].isdecimal() and i[-9:-5].isdecimal():
                    #print("92 line ",ff+"/"+i)
                    files.append(ff+"/"+i)
        for file in files:
            #print("95 line",file)
            if file[-4:]!="xlsx":
                continue
            try:
                wb=op.load_workbook(file)
                print(wb)
            except Exception as er:
                print("New workbook file :- "+"Continuing",er)
                continue
            try:
                wb.remove(wb['Sheet'])
            except Exception as err:
                print("New workbook file :- "+str(err))
            try:
                wb.remove(wb['Sheet1'])
            except Exception as err:
                print("New workbook file :- "+str(err))
            #print(readdata,type(readdata))
            for sheet in wb:
                #print("In loop sheet name",sheet)
                self.va=[]
                self.credit=[]
                #print(sheet["A2"].value)
                if sheet["A2"].value+".xlsx" in os.listdir(path1):
                    wd=op.load_workbook(path1+sheet["A2"].value+".xlsx")
                    try:
                        wp=wd[file.split("/")[-1][:-5]]
                    except:
                        wp=wd.create_sheet()
                        wp.title=file.split("/")[-1][:-5]
                        wp.append(["M/s","GSTIN No","Invoice No","Date","Description","Vehicle No.","Owner/Transporter","Quantity","Rate","Taxable Amount","Tax","Total Amount"])
                else:
                    wd=op.Workbook()
                    wp=wd.active
                    wp.title=file.split("/")[-1][:-5]
                    wp.append(["M/s","GSTIN No","Invoice No","Date","Description","Vehicle No.","Owner/Transporter","Quantity","Rate","Taxable Amount","Tax","Total Amount"])
                try:
                    readdata=xmlfile.readxmldata(path1+"invoicedetail.xml")
                    #print("New workbook file :- "+readdata,type(readdata))
                except:
                    print("New workbook file :- "+"readdata error")
                #print("New workbook file :- "+readdata,type(readdata))
                for row in range(4,sheet.max_row+1):
                    if row==4 and sheet[2][5].value not in readdata:
                        xmlfile.createxml(path1+"invoicedetail.xml","Invoice_No",datatag=[sheet[2][5].value])
                        wp.append([sheet[2][0].value,sheet[2][4].value,sheet[2][5].value,sheet[2][6].value,sheet[row][1].value,sheet[2][9].value,sheet[2][11].value,sheet[row][6].value,sheet[row][7].value,sheet[row][9].value,float(sheet[2][13].value+sheet[2][14].value+sheet[2][15].value+sheet[2][16].value+sheet[2][17].value),sheet[2][18].value])
                    elif sheet[2][5].value not in readdata:
                        wp.append([sheet[2][0].value,sheet[2][4].value,'','',sheet[row][1].value,'',sheet[2][11].value,sheet[row][6].value,sheet[row][7].value,sheet[row][9].value,'',''])
                    self.va.append([sheet["A2"].value,sheet["E2"].value,sheet[2][18].value,sheet["G2"].value])
                    self.credit.append(row-4)
                try:
                    print(wd)
                    wd.save(path1+sheet["A2"].value+".xlsx")
                except Exception as err:
                    print("New workbook file :- "+str(err),"125")
                try:
                    insert.inset(self)
                except Exception as err:
                    print("New workbook file :- "+"Error Happened"+str(err))
        for file in os.listdir(path1):
            if file[-4:]!="xlsx":
                continue
            wd=op.load_workbook(path1+file)
            ws=op.Workbook()
            wb=ws.active
            wb.title="Data"
            wb.append(["M/s","GSTIN No","Invoice No","Date","Description","Vehicle No.","Owner/Transporter","Quantity","Rate","Taxable Amount","Tax","Total Amount"])
            for sheet in wd:
                try:
                    [wb.append([j.value for j in sheet[i]]) for i in range(2,sheet.max_row+1)]
                except Exception as err:
                    print(str(err))
                    continue
            ws.save(path+file)
    def insertdata(self,master):
        master.dirr="C:/Users/"+os.getlogin()+"/Desktop/GST Data/"+merge.init()+"/"
        self.ddata=["Seller","Seller Address","Seller GST No.","Invoice NO.","Date","E-way Bill NO.","Vehicle No.","Owner/Transporter","Sr. No","Description of goods","Challan No","GST @","CESS @","TCS/TDS @","HSN no",
                    "Quantity","Rate","Taxable Amount","CGST","SGST","IGST","CESS","TCS/TDS","Grand Total"]
        path=master.dirr+master.date.get()[3:]+"/"
        if os.path.exists(path):
            pass
        else:
            os.makedirs(path)
        try:
            wb=op.load_workbook(master.dirr+master.date.get()[7:]+"input"+".xlsx")
        except:
            wb=op.Workbook()
        try:
            ws=wb[master.date.get()[3:]]
        except:
            ws=wb.create_sheet()
            ws.title=master.date.get()[3:]
            ws.append(self.ddata)
        total=0
        for i in range(len(master.data)):
            a=master.gstt[i]
            total=float(master.data[i][-1])+a[0]+a[1]+a[2]+a[3]+a[4]
        for i in range(len(master.data)):
            ws.append(master.detail+master.billdetail+master.data[i]+master.gstt[i]+[total])
        wb.save(master.dirr+master.date.get()[7:]+"input"+".xlsx")
        try:
            if os.path.exists(path):
                wb=op.load_workbook(path+master.date.get()[7:]+".xlsx")
            else:
                os.mkdir(path)
                wb=op.load_workbook(path+master.date.get()[7:]+".xlsx")
        except Exception as er:
            print("New workbook file :- "+str(er))
        try:
            wb=op.load_workbook(path+master.date.get()[3:]+"input"+".xlsx")
        except:
            wb=op.Workbook()
        try:
            ws=wb[master.date.get()[3:]]
        except:
            ws=wb.create_sheet()
            ws.title=master.date.get()[3:]
        ws.append(["GSTIN of Supplier","Invoice Number","Invoice date","Invoice Value","Place Of Supply","Reverse Charge",
                   "Invoice Type","Rate","Taxable Value ","Integrated Tax Paid","Central Tax Paid","State/UT Tax Paid",
                   "Cess Paid","Eligibility For ITC","Availed ITC Integrated Tax","Availed ITC Central Tax",
                   "Availed ITC State/UT Tax","Availed ITC Cess"])
        for i in range(len(master.data)):
            a=master.gstt[i]
            ws.append([master.detail[2],master.billdetail[0],master.billdetail[1],total,"",'','Regular',master.data[i][4],master.data[i][9],a[2],a[0],a[1],a[3],"Eligible",a[2],a[0],a[1],a[3]])
        try:
            wb.save(path+master.date.get()[3:]+"input"+".xlsx")
        except:
            print("New workbook file :- "+"Error")
